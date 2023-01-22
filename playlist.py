# libraries
import csv
import os
import re
import inspect
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import spotipy
from dotenv import load_dotenv
from dotenv import dotenv_values
from spotipy.oauth2 import SpotifyClientCredentials

# loading credentials from .env file
load_dotenv()
config0 = dotenv_values("keys.env").items()
config=list(config0)
CLIENT_ID = config[0][1]
CLIENT_SECRET = config[1][1]
OUTPUT_FILE_NAME = "track_info.csv"

# target playlist
PLAYLIST_LINK = config[2][1]


# authentication
client_credentials_manager = SpotifyClientCredentials(client_id=CLIENT_ID, client_secret=CLIENT_SECRET)

# creating spotify session object
session = spotipy.Spotify(client_credentials_manager=client_credentials_manager)

# getting uri from https link
if match := re.match(r"https://open.spotify.com/playlist/(.*)\?", PLAYLIST_LINK):
    playlist_uri = match.groups()[0]
else:
    raise ValueError("Expected format: https://open.spotify.com/playlist/...")

# getting list of tracks in a given playlist (note: max playlist length 100)
tracks = session.playlist_tracks(playlist_uri)["items"]

# creating csv file
with open(OUTPUT_FILE_NAME, "w", encoding="utf-8") as file:
    writer = csv.writer(file)
    
    # writing header column names
    writer.writerow(["Track", "Artist / Music band","Popularity (0 - 100)","Energy (0 - 1)","Tempo (60 - 180)","Length"])

    # extracting name and artist
    for track in tracks:
        name = track["track"]["name"]
        artists = ", ".join(
            [artist["name"] for artist in track["track"]["artists"]]
        )
        popularity=track["track"]["popularity"]
        
        energy=session.audio_features(track["track"]["id"])[0]["energy"]
        tempo=session.audio_features(track["track"]["id"])[0]["tempo"]
        duration_ms=session.audio_features(track["track"]["id"])[0]["duration_ms"]
        seconds=int((duration_ms/1000)%60)
        minutes=int((duration_ms/(1000*60))%60)
        length=("%d:%d" % (minutes, seconds))
       
        
        # write to csv
        writer.writerow([name, artists,popularity,energy,tempo,length])
 



# python file's directory working on it
module_path = inspect.getfile(inspect.currentframe())

# directory working on it
module_dir = os.path.realpath(os.path.dirname(module_path))       
 
module_path=module_dir+"/track_info.csv"

read_file = pd.read_csv (r'%s' % module_path)

#saving file as excel
excel_file_name="Track_Artist_Music_Band_Info.xlsx"
read_file.to_excel(excel_file_name, sheet_name="Spotify",index=False)

excel_file_path=module_dir+"/"+excel_file_name

wb = load_workbook(excel_file_path)
sheet = wb.active
# setting the font style to italic
sheet.cell(row = 1,column=1).font = Font(size = 13, italic = True,bold = True,color="191414")
sheet.cell(row = 1,column=2).font = Font(size = 13, italic = True,bold = True,color="191414")
sheet.cell(row = 1,column=3).font = Font(size = 13, italic = True,bold = True,color="191414")
sheet.cell(row = 1,column=4).font = Font(size = 13, italic = True,bold = True,color="191414")
sheet.cell(row = 1,column=5).font = Font(size = 13, italic = True,bold = True,color="191414")
sheet.cell(row = 1,column=6).font = Font(size = 13, italic = True,bold = True,color="191414")

sheet.cell(row = 1,column=1).fill  = PatternFill("solid", start_color="1DB954")
sheet.cell(row = 1,column=2).fill  = PatternFill("solid", start_color="1DB954")
sheet.cell(row = 1,column=3).fill  = PatternFill("solid", start_color="1DB954")
sheet.cell(row = 1,column=4).fill  = PatternFill("solid", start_color="1DB954")
sheet.cell(row = 1,column=5).fill  = PatternFill("solid", start_color="1DB954")
sheet.cell(row = 1,column=6).fill  = PatternFill("solid", start_color="1DB954")



sheet.cell(row = 1,column=1).border  = Border(bottom=Side(border_style='thick',color='000000'),
                                            right=Side(border_style='thick',color='000000'),
                                            left=Side(border_style='thick',color='000000'),
                                            top=Side(border_style='double',color='000000'))

sheet.cell(row = 1,column=2).border  = Border(right=Side(border_style='thick',color='000000'),
                                             bottom=Side(border_style='thick',color='000000'),
                                             top=Side(border_style='thick',color='000000'))

sheet.cell(row = 1,column=3).border  = Border(right=Side(border_style='thick',color='000000'),
                                             bottom=Side(border_style='thick',color='000000'),
                                             top=Side(border_style='thick',color='000000'))

sheet.cell(row = 1,column=4).border  = Border(right=Side(border_style='thick',color='000000'),
                                             bottom=Side(border_style='thick',color='000000'),
                                             top=Side(border_style='thick',color='000000'))
sheet.cell(row = 1,column=5).border  = Border(right=Side(border_style='thick',color='000000'),
                                             bottom=Side(border_style='thick',color='000000'),
                                             top=Side(border_style='thick',color='000000'))
sheet.cell(row = 1,column=6).border  = Border(right=Side(border_style='thick',color='000000'),
                                             bottom=Side(border_style='thick',color='000000'),
                                             top=Side(border_style='thick',color='000000'))


# column 1 settings
peach_c1 = '00FFCC99'
color_fill_c1='000000'
color_font_c1='191414'
for rows in sheet.iter_rows(min_row=2, max_row=len(tracks)+1,min_col=1, max_col=1): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=peach_c1, end_color=peach_c1,fill_type = "solid") 
            cell.border = Border(right=Side(border_style='thick',color=color_fill_c1),
                                                         bottom=Side(border_style='thick',color=color_fill_c1),
                                                         top=Side(border_style='thin',color=color_fill_c1))
            cell.font = Font(size = 13, italic = True,bold = False,color=color_font_c1)

# column 2 settings
peach_c2 = '00FFCC99'
color_fill_c2='000000'
color_font_c2='191414'
for rows in sheet.iter_rows(min_row=2, max_row=len(tracks)+1,min_col=2, max_col=2): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=peach_c2, end_color=peach_c2,fill_type = "solid") 
            cell.border = Border(right=Side(border_style='thick',color=color_fill_c2),
                                                         bottom=Side(border_style='thick',color=color_fill_c2),
                                                         top=Side(border_style='thin',color=color_fill_c2))
            cell.font = Font(size = 13, italic = True,bold = False,color=color_font_c2)

# column 3 settings
peach_c3 = '00FFCC99'
color_fill_c3='000000'
color_font_c3='191414'
for rows in sheet.iter_rows(min_row=2, max_row=len(tracks)+1,min_col=3, max_col=3): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=peach_c3, end_color=peach_c3,fill_type = "solid") 
            cell.border = Border(right=Side(border_style='thick',color=color_fill_c3),
                                                         bottom=Side(border_style='thick',color=color_fill_c3),
                                                         top=Side(border_style='thin',color=color_fill_c3))
            cell.font = Font(size = 13, italic = True,bold = False,color=color_font_c3)

# column 4 settings
peach_c4 = '00FFCC99'
color_fill_c4='000000'
color_font_c4='191414'
for rows in sheet.iter_rows(min_row=2, max_row=len(tracks)+1,min_col=4, max_col=4): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=peach_c4, end_color=peach_c4,fill_type = "solid") 
            cell.border = Border(right=Side(border_style='thick',color=color_fill_c4),
                                                         bottom=Side(border_style='thick',color=color_fill_c4),
                                                         top=Side(border_style='thin',color=color_fill_c4))
            cell.font = Font(size = 13, italic = True,bold = False,color=color_font_c4)

# column 5 settings
peach_c5 = '00FFCC99'
color_fill_c5='000000'
color_font_c5='191414'
for rows in sheet.iter_rows(min_row=2, max_row=len(tracks)+1,min_col=5, max_col=5): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=peach_c5, end_color=peach_c5,fill_type = "solid") 
            cell.border = Border(right=Side(border_style='thick',color=color_fill_c5),
                                                         bottom=Side(border_style='thick',color=color_fill_c5),
                                                         top=Side(border_style='thin',color=color_fill_c5))
            cell.font = Font(size = 13, italic = True,bold = False,color=color_font_c5)
            

# column 6 settings
peach_c6 = '00FFCC99'
color_fill_c6='000000'
color_font_c6='191414'
for rows in sheet.iter_rows(min_row=2, max_row=len(tracks)+1,min_col=6, max_col=6): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=peach_c6, end_color=peach_c6,fill_type = "solid") 
            cell.border = Border(right=Side(border_style='thick',color=color_fill_c6),
                                                         bottom=Side(border_style='thick',color=color_fill_c6),
                                                         top=Side(border_style='thin',color=color_fill_c6))
            cell.font = Font(size = 13, italic = True,bold = False,color=color_font_c6)
          
 
# alignment for C column          
for i in range(len(tracks)+1):
     if i!=0:
      value="C"+str(i+1)
      sheet[value].alignment = Alignment(horizontal="center")    
 
    
# alignment for D column          
for i in range(len(tracks)+1):
     if i!=0:
      value="D"+str(i+1)
      sheet[value].alignment = Alignment(horizontal="center")


     
# alignment for E column          
for i in range(len(tracks)+1):
     if i!=0:
      value="E"+str(i+1)
      sheet[value].alignment = Alignment(horizontal="center")

# alignment for F column          
for i in range(len(tracks)+1):
     if i!=0:
      value="F"+str(i+1)
      sheet[value].alignment = Alignment(horizontal="center")
    

#column width settings
sheet.column_dimensions['A'].width = 28
sheet.column_dimensions['B'].width = 28
sheet.column_dimensions['C'].width = 28
sheet.column_dimensions['D'].width = 28
sheet.column_dimensions['E'].width = 28
sheet.column_dimensions['F'].width = 28



#resaving the file
wb.save("Track_Artist_Music_Band_Info.xlsx")